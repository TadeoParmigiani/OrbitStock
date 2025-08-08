import os
from datetime import datetime, date
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, Http404, JsonResponse
from django.conf import settings
from django.db.models import Sum, Count, F
from django.utils import timezone
from users.decorators import role_required
from .models import Report
from sales.models import Venta, DetalleVenta
from products.models import Producto

# Importaciones para generar archivos
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

@role_required('admin')
def reports_list(request, tipo):
    """Vista principal de reportes según tipo"""
    if tipo not in ['ventas', 'stock']:
        raise Http404("Tipo de reporte inválido.")
    
    reports_history = Report.objects.filter(tipo=tipo).order_by('-fecha_creacion')[:10]
    template = 'reportsSales.html' if tipo == 'ventas' else 'reportsStock.html'
    
    # Debug: imprimir información
    print(f"Tipo de reporte: {tipo}")
    print(f"Cantidad de reportes encontrados: {reports_history.count()}")
    for report in reports_history:
        print(f"Reporte ID: {report.id}, Estado: {report.estado}, Archivo: {report.ruta_archivo}")
    
    return render(request, template, {
        'reports_history': reports_history,
        'tipo': tipo,
    })

@role_required('admin')
def create_report(request):
    """Crear un nuevo reporte"""
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        formato = request.POST.get('formato')
        descripcion = request.POST.get('descripcion', '')
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        
        print(f"=== CREAR REPORTE ===")
        print(f"Datos recibidos - Tipo: {tipo}, Formato: {formato}")
        print(f"Fechas recibidas: '{fecha_inicio}' a '{fecha_fin}'")
        
        try:
            # Validar y convertir fechas
            if not fecha_inicio or not fecha_fin:
                messages.error(request, 'Debe seleccionar fecha de inicio y fecha de fin')
                return redirect('reportsList', tipo=tipo)
            
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            
            print(f"Fechas convertidas: {fecha_inicio_obj} a {fecha_fin_obj}")
            
            if fecha_inicio_obj > fecha_fin_obj:
                messages.error(request, 'La fecha de inicio no puede ser mayor que la fecha de fin')
                return redirect('reportsList', tipo=tipo)
            
            # Verificar que hay datos en el rango antes de crear el reporte
            print(f"=== VERIFICACIÓN PREVIA DE DATOS ===")
            if tipo == 'ventas':
                ventas_test = Venta.objects.filter(
                    fecha__date__gte=fecha_inicio_obj,
                    fecha__date__lte=fecha_fin_obj
                ).count()
                print(f"Ventas en el rango seleccionado: {ventas_test}")
                
                if ventas_test == 0:
                    # Verificar ventas en general
                    todas_ventas = Venta.objects.count()
                    print(f"Total de ventas en BD: {todas_ventas}")
                    
                    if todas_ventas > 0:
                        # Mostrar fechas disponibles
                        fechas_disponibles = Venta.objects.values_list('fecha__date', flat=True).distinct().order_by('fecha__date')
                        print(f"Fechas de ventas disponibles: {list(fechas_disponibles)}")
                        
                        messages.warning(request, f'No se encontraron ventas entre {fecha_inicio_obj} y {fecha_fin_obj}. Fechas disponibles: {", ".join(str(f) for f in fechas_disponibles[:5])}')
                    else:
                        messages.error(request, 'No hay ventas registradas en el sistema')
                        return redirect('reportsList', tipo=tipo)
            
            elif tipo == 'stock':
                productos_test = DetalleVenta.objects.filter(
                    venta__fecha__date__gte=fecha_inicio_obj,
                    venta__fecha__date__lte=fecha_fin_obj
                ).count()
                print(f"Detalles de venta en el rango seleccionado: {productos_test}")
                
                if productos_test == 0:
                    todos_detalles = DetalleVenta.objects.count()
                    print(f"Total de detalles de venta en BD: {todos_detalles}")
                    
                    if todos_detalles > 0:
                        fechas_disponibles = DetalleVenta.objects.values_list('venta__fecha__date', flat=True).distinct().order_by('venta__fecha__date')
                        print(f"Fechas de ventas con productos disponibles: {list(fechas_disponibles)}")
                        
                        messages.warning(request, f'No se encontraron productos vendidos entre {fecha_inicio_obj} y {fecha_fin_obj}. Fechas disponibles: {", ".join(str(f) for f in fechas_disponibles[:5])}')
                    else:
                        messages.error(request, 'No hay detalles de venta registrados en el sistema')
                        return redirect('reportsList', tipo=tipo)
            
            # Crear registro de reporte
            report = Report.objects.create(
                tipo=tipo,
                formato=formato,
                descripcion=descripcion,
                fecha_inicio=fecha_inicio_obj,
                fecha_fin=fecha_fin_obj,
                creado_por=request.user,
                estado='pendiente'
            )
            
            print(f"Reporte creado con ID: {report.id}")
            
            # Generar nombre de archivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            extension = 'pdf' if formato == 'pdf' else 'xlsx'
            filename = f'reporte_{tipo}_{timestamp}.{extension}'
            reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
            
            # Crear directorio si no existe
            if not os.path.exists(reports_dir):
                os.makedirs(reports_dir)
                print(f"Directorio creado: {reports_dir}")
            
            file_path = os.path.join(reports_dir, filename)
            print(f"Ruta del archivo: {file_path}")
            
            # Generar el reporte según el tipo
            if tipo == 'ventas':
                success = generate_sales_report(report, file_path, fecha_inicio_obj, fecha_fin_obj, formato)
            elif tipo == 'stock':
                success = generate_stock_report(report, file_path, fecha_inicio_obj, fecha_fin_obj, formato)
            else:
                success = False
            
            if success:
                # Actualizar registro con ruta relativa para MEDIA_URL
                relative_path = f'reports/{filename}'
                report.ruta_archivo = relative_path
                if os.path.exists(file_path):
                    report.tamaño_archivo = os.path.getsize(file_path)
                report.estado = 'completado'
                report.save()
                
                print(f"Reporte completado. Tamaño: {report.tamaño_archivo} bytes")
                
                # CAMBIO PRINCIPAL: Agregar parámetro para indicar descarga automática
                messages.success(request, f'Reporte {report.get_tipo_display()} creado exitosamente')
                
                # Redirigir al historial con parámetro para descarga automática
                from django.urls import reverse
                url = reverse('reportsList', kwargs={'tipo': tipo}) + f'?download={report.id}'
                return redirect(url)
                
            else:
                report.estado = 'fallido'
                report.save()
                messages.error(request, 'Error al generar el reporte')
                
        except Exception as e:
            print(f"Error en create_report: {str(e)}")
            import traceback
            print(traceback.format_exc())
            if 'report' in locals():
                report.estado = 'fallido'
                report.save()
            messages.error(request, f'Error al crear reporte: {str(e)}')
            # Si no tenemos tipo, redirigir a una página por defecto
            tipo = request.POST.get('tipo', 'ventas')  # Default a ventas si no hay tipo
            return redirect('reportsList', tipo=tipo)
    
    # Si es GET request, redirigir a reportes de ventas por defecto
    return redirect('reportsList', tipo='ventas')

def generate_sales_report(report, file_path, fecha_inicio, fecha_fin, formato):
    """Generar reporte de ventas"""
    try:
        print(f"=== DEBUG REPORTE DE VENTAS ===")
        print(f"Fechas de búsqueda: {fecha_inicio} a {fecha_fin}")
        print(f"Tipos de fechas: {type(fecha_inicio)} - {type(fecha_fin)}")
        
        # Primero verificar todas las ventas disponibles
        todas_ventas = Venta.objects.all()
        print(f"Total de ventas en BD: {todas_ventas.count()}")
        
        if todas_ventas.exists():
            print("Primeras 3 ventas en BD:")
            for venta in todas_ventas[:3]:
                fecha_venta = venta.fecha.date() if hasattr(venta.fecha, 'date') else venta.fecha
                print(f"  - ID: {venta.id}, Fecha: {venta.fecha} (solo fecha: {fecha_venta}), Total: {venta.total}")
        
        # Intentar diferentes filtros para encontrar las ventas
        print("\n=== PROBANDO DIFERENTES FILTROS ===")
        
        # Filtro 1: Rango de fechas con date()
        ventas1 = Venta.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin
        ).select_related('cliente').order_by('fecha')
        print(f"Filtro 1 (fecha__date__gte/lte): {ventas1.count()} ventas")
        
        # Filtro 2: Rango de fechas con range
        from datetime import datetime, time
        fecha_inicio_dt = datetime.combine(fecha_inicio, time.min)
        fecha_fin_dt = datetime.combine(fecha_fin, time.max)
        
        ventas2 = Venta.objects.filter(
            fecha__range=(fecha_inicio_dt, fecha_fin_dt)
        ).select_related('cliente').order_by('fecha')
        print(f"Filtro 2 (fecha__range): {ventas2.count()} ventas")
        
        # Filtro 3: Solo por fecha usando contains
        ventas3 = Venta.objects.filter(
            fecha__date__range=[fecha_inicio, fecha_fin]
        ).select_related('cliente').order_by('fecha')
        print(f"Filtro 3 (fecha__date__range): {ventas3.count()} ventas")
        
        # Usar el filtro que encuentre más ventas
        ventas_opciones = [
            ("Filtro 1", ventas1),
            ("Filtro 2", ventas2), 
            ("Filtro 3", ventas3)
        ]
        
        # Elegir el que tenga más resultados
        mejor_opcion = max(ventas_opciones, key=lambda x: x[1].count())
        ventas = mejor_opcion[1]
        
        print(f"\nUsando {mejor_opcion[0]} con {ventas.count()} ventas encontradas")
        
        # Si aún no encuentra ventas, expandir el rango
        if ventas.count() == 0:
            print("\nNo se encontraron ventas en el rango, expandiendo búsqueda...")
            
            # Buscar ventas en los últimos 30 días
            from datetime import timedelta
            fecha_expandida = fecha_fin - timedelta(days=30)
            ventas_expandidas = Venta.objects.filter(
                fecha__date__gte=fecha_expandida
            ).select_related('cliente').order_by('fecha')
            
            print(f"Ventas en los últimos 30 días: {ventas_expandidas.count()}")
            
            if ventas_expandidas.count() > 0:
                print("Fechas de ventas disponibles:")
                fechas_disponibles = ventas_expandidas.values_list('fecha__date', flat=True).distinct()
                for fecha in fechas_disponibles[:10]:
                    print(f"  - {fecha}")
        
        # Debug: mostrar ventas encontradas
        if ventas.exists():
            print(f"\nVentas encontradas para el reporte:")
            for venta in ventas[:5]:
                print(f"  - ID: {venta.id}, Fecha: {venta.fecha}, Total: {venta.total}")
        else:
            print("\n¡ATENCIÓN! No se encontraron ventas para el reporte")
        
        if formato == 'pdf':
            return generate_sales_pdf(report, file_path, ventas, fecha_inicio, fecha_fin)
        else:
            return generate_sales_excel(report, file_path, ventas, fecha_inicio, fecha_fin)
            
    except Exception as e:
        print(f"Error generando reporte de ventas: {e}")
        import traceback
        print(traceback.format_exc())
        return False

def generate_stock_report(report, file_path, fecha_inicio, fecha_fin, formato):
    """Generar reporte de stock/productos vendidos"""
    try:
        print(f"=== DEBUG REPORTE DE STOCK ===")
        print(f"Fechas de búsqueda: {fecha_inicio} a {fecha_fin}")
        
        # Primero verificar todos los detalles disponibles
        todos_detalles = DetalleVenta.objects.all()
        print(f"Total de detalles en BD: {todos_detalles.count()}")
        
        if todos_detalles.exists():
            print("Primeros 3 detalles en BD:")
            for detalle in todos_detalles[:3]:
                fecha_venta = detalle.venta.fecha.date() if hasattr(detalle.venta.fecha, 'date') else detalle.venta.fecha
                print(f"  - Producto: {detalle.producto.nombre}, Fecha venta: {detalle.venta.fecha} (solo fecha: {fecha_venta}), Cantidad: {detalle.cantidad}")
        
        # Intentar diferentes filtros
        print("\n=== PROBANDO DIFERENTES FILTROS ===")
        
        # Filtro 1: Con date()
        productos1 = DetalleVenta.objects.filter(
            venta__fecha__date__gte=fecha_inicio,
            venta__fecha__date__lte=fecha_fin
        ).select_related('producto', 'producto__categoria', 'venta').values(
            'producto__nombre',
            'producto__codigo', 
            'producto__categoria__nombre'
        ).annotate(
            cantidad_vendida=Sum('cantidad'),
            monto_total=Sum(F('precio_unitario') * F('cantidad'))
        ).order_by('-cantidad_vendida')
        print(f"Filtro 1 (fecha__date__gte/lte): {productos1.count()} productos")
        
        # Filtro 2: Con range
        from datetime import datetime, time
        fecha_inicio_dt = datetime.combine(fecha_inicio, time.min)
        fecha_fin_dt = datetime.combine(fecha_fin, time.max)
        
        productos2 = DetalleVenta.objects.filter(
            venta__fecha__range=(fecha_inicio_dt, fecha_fin_dt)
        ).select_related('producto', 'producto__categoria', 'venta').values(
            'producto__nombre',
            'producto__codigo', 
            'producto__categoria__nombre'
        ).annotate(
            cantidad_vendida=Sum('cantidad'),
            monto_total=Sum(F('precio_unitario') * F('cantidad'))
        ).order_by('-cantidad_vendida')
        print(f"Filtro 2 (fecha__range): {productos2.count()} productos")
        
        # Filtro 3: Con date__range
        productos3 = DetalleVenta.objects.filter(
            venta__fecha__date__range=[fecha_inicio, fecha_fin]
        ).select_related('producto', 'producto__categoria', 'venta').values(
            'producto__nombre',
            'producto__codigo', 
            'producto__categoria__nombre'
        ).annotate(
            cantidad_vendida=Sum('cantidad'),
            monto_total=Sum(F('precio_unitario') * F('cantidad'))
        ).order_by('-cantidad_vendida')
        print(f"Filtro 3 (fecha__date__range): {productos3.count()} productos")
        
        # Usar el filtro que encuentre más productos
        productos_opciones = [
            ("Filtro 1", productos1),
            ("Filtro 2", productos2), 
            ("Filtro 3", productos3)
        ]
        
        mejor_opcion = max(productos_opciones, key=lambda x: x[1].count())
        productos_vendidos = mejor_opcion[1]
        
        print(f"\nUsando {mejor_opcion[0]} con {productos_vendidos.count()} productos encontrados")
        
        # Si no encuentra productos, expandir búsqueda
        if productos_vendidos.count() == 0:
            print("\nNo se encontraron productos, expandiendo búsqueda...")
            
            from datetime import timedelta
            fecha_expandida = fecha_fin - timedelta(days=30)
            productos_expandidos = DetalleVenta.objects.filter(
                venta__fecha__date__gte=fecha_expandida
            ).select_related('producto', 'producto__categoria', 'venta').values(
                'producto__nombre',
                'producto__codigo', 
                'producto__categoria__nombre'
            ).annotate(
                cantidad_vendida=Sum('cantidad'),
                monto_total=Sum(F('precio_unitario') * F('cantidad'))
            ).order_by('-cantidad_vendida')
            
            print(f"Productos en los últimos 30 días: {productos_expandidos.count()}")
        
        # Debug: mostrar productos encontrados
        if productos_vendidos.exists():
            print(f"\nProductos encontrados para el reporte:")
            for producto in productos_vendidos[:3]:
                print(f"  - {producto}")
        else:
            print("\n¡ATENCIÓN! No se encontraron productos para el reporte")
        
        if formato == 'pdf':
            return generate_stock_pdf(report, file_path, productos_vendidos, fecha_inicio, fecha_fin)
        else:
            return generate_stock_excel(report, file_path, productos_vendidos, fecha_inicio, fecha_fin)
            
    except Exception as e:
        print(f"Error generando reporte de stock: {e}")
        import traceback
        print(traceback.format_exc())
        return False

def generate_sales_pdf(report, file_path, ventas, fecha_inicio, fecha_fin):
    """Generar PDF de reporte de ventas con estética mejorada"""
    try:
        doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=60)
        styles = getSampleStyleSheet()
        story = []
        
        # Título principal con estilo mejorado
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a365d'),
            spaceAfter=20,
            spaceBefore=10,
            alignment=1,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("📊 REPORTE DE VENTAS", title_style))
        
        # Línea decorativa
        story.append(HRFlowable(width="100%", thickness=3, color=colors.HexColor('#3182ce')))
        story.append(Spacer(1, 15))
        
        # Período con mejor estilo
        period_style = ParagraphStyle(
            'Period',
            parent=styles['Normal'],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor('#4a5568'),
            fontName='Helvetica-Bold',
            spaceAfter=25
        )
        story.append(Paragraph(f"📅 Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}", period_style))
        
        # Resumen con cards estilizados
        total_ventas = ventas.count()
        total_general = sum(venta.total for venta in ventas) if ventas.exists() else Decimal('0')
        
        summary_data = [
            ['📈 TOTAL VENTAS', '💰 MONTO TOTAL'],
            [f'{total_ventas} ventas', f'${total_general:,.2f}']
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4299e1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 15),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 15),
            
            # Valores
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ebf8ff')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#2d3748')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 16),
            ('TOPPADDING', (0, 1), (-1, 1), 20),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 20),
            
            # Bordes
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#4299e1')),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.HexColor('#bee3f8')),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Subtítulo para la tabla
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=16,
            textColor=colors.HexColor('#2d3748'),
            fontName='Helvetica-Bold',
            spaceAfter=15,
            alignment=1
        )
        story.append(Paragraph("🧾 DETALLE DE VENTAS", subtitle_style))
        
        # Datos de la tabla con colores mejorados
        data = [['#', 'Fecha y Hora', 'Cliente', 'Método Pago', 'Total']]
        
        if ventas.exists():
            for i, venta in enumerate(ventas, 1):
                cliente = venta.cliente.nombre if venta.cliente else 'Cliente General'
                data.append([
                    str(i),
                    venta.fecha.strftime('%d/%m/%Y\n%H:%M'),
                    cliente,
                    venta.get_metodo_pago_display() if hasattr(venta, 'get_metodo_pago_display') else venta.metodo_pago,
                    f"${venta.total:,.2f}"
                ])
        else:
            data.append(['', 'No se encontraron ventas en el período seleccionado', '', '', '$0.00'])
        
        # Fila de total con estilo destacado
        data.append(['', '', '', '💎 TOTAL GENERAL', f"${total_general:,.2f}"])
        
        # Crear tabla con mejor diseño
        table = Table(data, colWidths=[0.6*inch, 1.4*inch, 2.2*inch, 1.4*inch, 1.2*inch])
        table.setStyle(TableStyle([
            # Encabezado principal
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b6cb0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Filas de datos con colores alternados
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f7fafc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#edf2f7')]),
            ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#2d3748')),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('TOPPADDING', (0, 1), (-1, -2), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
            
            # Fila de total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#38a169')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('TOPPADDING', (0, -1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
            
            # Bordes y grillas
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#4a5568')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
        ]))
        
        story.append(table)
        doc.build(story)
        
        # Actualizar estadísticas del reporte
        report.total_registros = total_ventas
        report.monto_total = total_general
        report.save()
        
        print(f"PDF generado exitosamente. Registros: {total_ventas}, Total: ${total_general}")
        
        return True
    except Exception as e:
        print(f"Error generando PDF de ventas: {e}")
        import traceback
        print(traceback.format_exc())
        return False

def generate_sales_excel(report, file_path, ventas, fecha_inicio, fecha_fin):
    """Generar Excel de reporte de ventas con estética mejorada"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "📊 Reporte de Ventas"
        
        # Estilos mejorados
        # Título principal
        title_font = Font(bold=True, size=18, color="1A365D")
        title_fill = PatternFill("solid", fgColor="E6FFFA")
        
        # Encabezados
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2B6CB0")
        
        # Totales
        total_font = Font(bold=True, color="FFFFFF", size=12)
        total_fill = PatternFill("solid", fgColor="38A169")
        
        # Resumen
        summary_font = Font(bold=True, size=14, color="2D3748")
        summary_fill = PatternFill("solid", fgColor="EBF8FF")
        
        # Bordes
        thin_border = Border(
            left=Side(border_style="thin", color="CBD5E0"),
            right=Side(border_style="thin", color="CBD5E0"),
            top=Side(border_style="thin", color="CBD5E0"),
            bottom=Side(border_style="thin", color="CBD5E0")
        )
        
        thick_border = Border(
            left=Side(border_style="thick", color="4A5568"),
            right=Side(border_style="thick", color="4A5568"),
            top=Side(border_style="thick", color="4A5568"),
            bottom=Side(border_style="thick", color="4A5568")
        )
        
        # Título y período
        ws.merge_cells('A1:E3')
        title_cell = ws['A1']
        title_cell.value = "📊 REPORTE DE VENTAS"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = thick_border
        
        ws.merge_cells('A4:E4')
        period_cell = ws['A4']
        period_cell.value = f"📅 Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        period_cell.font = Font(bold=True, size=12, color="4A5568")
        period_cell.alignment = Alignment(horizontal="center")
        
        # Resumen con mejor diseño
        total_ventas = ventas.count()
        total_general = sum(venta.total for venta in ventas) if ventas.exists() else Decimal('0')
        
        # Cards de resumen
        ws.merge_cells('A6:B6')
        ws['A6'] = "📈 TOTAL DE VENTAS"
        ws['A6'].font = summary_font
        ws['A6'].fill = summary_fill
        ws['A6'].alignment = Alignment(horizontal="center")
        ws['A6'].border = thin_border
        
        ws.merge_cells('D6:E6')
        ws['D6'] = "💰 MONTO TOTAL"
        ws['D6'].font = summary_font
        ws['D6'].fill = summary_fill
        ws['D6'].alignment = Alignment(horizontal="center")
        ws['D6'].border = thin_border
        
        ws.merge_cells('A7:B7')
        ws['A7'] = f"{total_ventas} ventas"
        ws['A7'].font = Font(bold=True, size=16, color="2B6CB0")
        ws['A7'].alignment = Alignment(horizontal="center")
        ws['A7'].border = thin_border
        
        ws.merge_cells('D7:E7')
        ws['D7'] = f"${total_general:,.2f}"
        ws['D7'].font = Font(bold=True, size=16, color="38A169")
        ws['D7'].alignment = Alignment(horizontal="center")
        ws['D7'].border = thin_border
        
        # Encabezados de tabla
        headers = ['#', 'Fecha y Hora', 'Cliente', 'Método Pago', 'Total']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
        
        # Datos con colores alternados
        row = 11
        
        if ventas.exists():
            for i, venta in enumerate(ventas, 1):
                cliente = venta.cliente.nombre if venta.cliente else 'Cliente General'
                
                # Color de fila alternado
                row_fill = PatternFill("solid", fgColor="F7FAFC") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                
                cells_data = [
                    (i, "center"),
                    (venta.fecha.strftime('%d/%m/%Y\n%H:%M'), "center"),
                    (cliente, "left"),
                    (venta.get_metodo_pago_display() if hasattr(venta, 'get_metodo_pago_display') else venta.metodo_pago, "center"),
                    (f"${venta.total:,.2f}", "right")
                ]
                
                for col, (value, align) in enumerate(cells_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.fill = row_fill
                    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
                    cell.border = thin_border
                    cell.font = Font(size=10, color="2D3748")
                
                row += 1
        else:
            no_data_fill = PatternFill("solid", fgColor="FED7D7")
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "No se encontraron ventas en el período seleccionado"
            cell.fill = no_data_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.cell(row=row, column=5, value="$0.00").border = thin_border
            row += 1
        
        # Fila de total
        ws.merge_cells(f'A{row}:D{row}')
        total_label = ws[f'A{row}']
        total_label.value = "💎 TOTAL GENERAL"
        total_label.font = total_font
        total_label.fill = total_fill
        total_label.alignment = Alignment(horizontal="center", vertical="center")
        total_label.border = thick_border
        
        total_cell = ws.cell(row=row, column=5, value=f"${total_general:,.2f}")
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.border = thick_border
        
        # Ajustar ancho de columnas
        column_widths = [6, 18, 30, 18, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Ajustar altura de filas
        ws.row_dimensions[1].height = 45
        ws.row_dimensions[6].height = 25
        ws.row_dimensions[7].height = 30
        ws.row_dimensions[10].height = 20
        
        wb.save(file_path)
        
        # Actualizar estadísticas del reporte
        report.total_registros = total_ventas
        report.monto_total = total_general
        report.save()
        
        print(f"Excel generado exitosamente. Registros: {total_ventas}, Total: ${total_general}")
        
        return True
    except Exception as e:
        print(f"Error generando Excel de ventas: {e}")
        import traceback
        print(traceback.format_exc())
        return False

def generate_stock_pdf(report, file_path, productos_vendidos, fecha_inicio, fecha_fin):
    """Generar PDF de reporte de stock con estética mejorada"""
    try:
        doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=60, bottomMargin=60)
        styles = getSampleStyleSheet()
        story = []
        
        # Título principal
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a202c'),
            spaceAfter=20,
            spaceBefore=10,
            alignment=1,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph("📦 REPORTE DE STOCK VENDIDO", title_style))
        
        # Línea decorativa
        story.append(HRFlowable(width="100%", thickness=3, color=colors.HexColor('#38a169')))
        story.append(Spacer(1, 15))
        
        # Período
        period_style = ParagraphStyle(
            'Period',
            parent=styles['Normal'],
            fontSize=14,
            alignment=1,
            textColor=colors.HexColor('#4a5568'),
            fontName='Helvetica-Bold',
            spaceAfter=25
        )
        story.append(Paragraph(f"📅 Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}", period_style))
        
        # Resumen con cards
        total_productos = productos_vendidos.count()
        total_cantidad = sum(item['cantidad_vendida'] or 0 for item in productos_vendidos)
        total_general = sum(item['monto_total'] or Decimal('0') for item in productos_vendidos)
        
        summary_data = [
            ['🏷️ PRODUCTOS', '📊 UNIDADES', '💰 TOTAL'],
            [f'{total_productos} productos', f'{total_cantidad} unidades', f'${total_general:,.2f}']
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38a169')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Valores
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f0fff4')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1a202c')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('TOPPADDING', (0, 1), (-1, 1), 15),
            ('BOTTOMPADDING', (0, 1), (-1, 1), 15),
            
            # Bordes
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#38a169')),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.HexColor('#9ae6b4')),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Subtítulo
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=16,
            textColor=colors.HexColor('#2d3748'),
            fontName='Helvetica-Bold',
            spaceAfter=15,
            alignment=1
        )
        story.append(Paragraph("📋 DETALLE DE PRODUCTOS VENDIDOS", subtitle_style))
        
        # Datos de la tabla
        data = [['#', 'Producto', 'Código', 'Categoría', 'Cant.', 'Total' ]]
        
        if productos_vendidos.exists():
            for i, item in enumerate(productos_vendidos, 1):
                data.append([
                    str(i),
                    item['producto__nombre'] or 'N/A',
                    item['producto__codigo'] or 'N/A',
                    item['producto__categoria__nombre'] or 'Sin categoría',
                    f"{item['cantidad_vendida'] or 0:,}",
                    f"${(item['monto_total'] or Decimal('0')):,.2f}"
                ])
        else:
            data.append(['', 'No se encontraron productos vendidos', '', '', '0', '$0.00'])
        
        # Agregar total
        data.append(['', '', '', '💎 TOTALES', f'{total_cantidad:,}', f"${total_general:,.2f}"])
        
        # Crear tabla
        table = Table(data, colWidths=[0.5*inch, 2.2*inch, 1*inch, 1.5*inch, 0.8*inch, 1.2*inch])
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f855a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Filas de datos
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#f7fafc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#edf2f7')]),
            ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#2d3748')),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('TOPPADDING', (0, 1), (-1, -2), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 8),
            
            # Fila de total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#3182ce')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('TOPPADDING', (0, -1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
            
            # Bordes
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#4a5568')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
        ]))
        
        story.append(table)
        doc.build(story)
        
        # Actualizar estadísticas del reporte
        report.total_registros = total_productos
        report.monto_total = total_general
        report.save()
        
        print(f"PDF de stock generado exitosamente. Registros: {total_productos}, Total: ${total_general}")
        
        return True
    except Exception as e:
        print(f"Error generando PDF de stock: {e}")
        import traceback
        print(traceback.format_exc())
        return False

def generate_stock_excel(report, file_path, productos_vendidos, fecha_inicio, fecha_fin):
    """Generar Excel de reporte de stock con estética mejorada"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "📦 Reporte de Stock"
        
        # Estilos
        title_font = Font(bold=True, size=18, color="1A202C")
        title_fill = PatternFill("solid", fgColor="F0FFF4")
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2F855A")
        
        total_font = Font(bold=True, color="FFFFFF", size=12)
        total_fill = PatternFill("solid", fgColor="3182CE")
        
        summary_font = Font(bold=True, size=14, color="2D3748")
        summary_fill = PatternFill("solid", fgColor="F0FFF4")
        
        thin_border = Border(
            left=Side(border_style="thin", color="CBD5E0"),
            right=Side(border_style="thin", color="CBD5E0"),
            top=Side(border_style="thin", color="CBD5E0"),
            bottom=Side(border_style="thin", color="CBD5E0")
        )
        
        thick_border = Border(
            left=Side(border_style="thick", color="4A5568"),
            right=Side(border_style="thick", color="4A5568"),
            top=Side(border_style="thick", color="4A5568"),
            bottom=Side(border_style="thick", color="4A5568")
        )
        
        # Título
        ws.merge_cells('A1:F3')
        title_cell = ws['A1']
        title_cell.value = "📦 REPORTE DE STOCK VENDIDO"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = thick_border
        
        # Período
        ws.merge_cells('A4:F4')
        period_cell = ws['A4']
        period_cell.value = f"📅 Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        period_cell.font = Font(bold=True, size=12, color="4A5568")
        period_cell.alignment = Alignment(horizontal="center")
        
        # Resumen
        total_productos = productos_vendidos.count()
        total_cantidad = sum(item['cantidad_vendida'] or 0 for item in productos_vendidos)
        total_general = sum(item['monto_total'] or Decimal('0') for item in productos_vendidos)
        
        # Cards de resumen
        ws.merge_cells('A6:B6')
        ws['A6'] = "🏷️ PRODUCTOS"
        ws['A6'].font = summary_font
        ws['A6'].fill = summary_fill
        ws['A6'].alignment = Alignment(horizontal="center")
        ws['A6'].border = thin_border
        
        ws.merge_cells('C6:D6')
        ws['C6'] = "📊 UNIDADES"
        ws['C6'].font = summary_font
        ws['C6'].fill = summary_fill
        ws['C6'].alignment = Alignment(horizontal="center")
        ws['C6'].border = thin_border
        
        ws.merge_cells('E6:F6')
        ws['E6'] = "💰 TOTAL"
        ws['E6'].font = summary_font
        ws['E6'].fill = summary_fill
        ws['E6'].alignment = Alignment(horizontal="center")
        ws['E6'].border = thin_border
        
        ws.merge_cells('A7:B7')
        ws['A7'] = f"{total_productos} productos"
        ws['A7'].font = Font(bold=True, size=14, color="2F855A")
        ws['A7'].alignment = Alignment(horizontal="center")
        ws['A7'].border = thin_border
        
        ws.merge_cells('C7:D7')
        ws['C7'] = f"{total_cantidad:,} unidades"
        ws['C7'].font = Font(bold=True, size=14, color="3182CE")
        ws['C7'].alignment = Alignment(horizontal="center")
        ws['C7'].border = thin_border
        
        ws.merge_cells('E7:F7')
        ws['E7'] = f"${total_general:,.2f}"
        ws['E7'].font = Font(bold=True, size=14, color="38A169")
        ws['E7'].alignment = Alignment(horizontal="center")
        ws['E7'].border = thin_border
        
        # Encabezados de tabla
        headers = ['#', 'Producto', 'Código', 'Categoría', 'Cantidad', 'Total' ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border
        
        # Datos con colores alternados
        row = 11
        
        if productos_vendidos.exists():
            for i, item in enumerate(productos_vendidos, 1):
                # Color de fila alternado
                row_fill = PatternFill("solid", fgColor="F7FAFC") if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                
                cells_data = [
                    (i, "center"),
                    (item['producto__nombre'] or 'N/A', "left"),
                    (item['producto__codigo'] or 'N/A', "center"),
                    (item['producto__categoria__nombre'] or 'Sin categoría', "center"),
                    (f"{item['cantidad_vendida'] or 0:,}", "right"),
                    (f"${(item['monto_total'] or Decimal('0')):,.2f}", "right")
                ]
                
                for col, (value, align) in enumerate(cells_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.fill = row_fill
                    cell.alignment = Alignment(horizontal=align, vertical="center")
                    cell.border = thin_border
                    cell.font = Font(size=10, color="2D3748")
                
                row += 1
        else:
            no_data_fill = PatternFill("solid", fgColor="FED7D7")
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = "No se encontraron productos vendidos"
            cell.fill = no_data_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.cell(row=row, column=6, value="$0.00").border = thin_border
            row += 1
        
        # Fila de totales
        ws.merge_cells(f'A{row}:D{row}')
        total_label = ws[f'A{row}']
        total_label.value = "💎 TOTALES GENERALES"
        total_label.font = total_font
        total_label.fill = total_fill
        total_label.alignment = Alignment(horizontal="center", vertical="center")
        total_label.border = thick_border
        
        # Cantidad total
        cantidad_cell = ws.cell(row=row, column=5, value=f"{total_cantidad:,}")
        cantidad_cell.font = total_font
        cantidad_cell.fill = total_fill
        cantidad_cell.alignment = Alignment(horizontal="right", vertical="center")
        cantidad_cell.border = thick_border
        
        # Monto total
        monto_cell = ws.cell(row=row, column=6, value=f"${total_general:,.2f}")
        monto_cell.font = total_font
        monto_cell.fill = total_fill
        monto_cell.alignment = Alignment(horizontal="right", vertical="center")
        monto_cell.border = thick_border
        
        # Ajustar ancho de columnas
        column_widths = [6, 35, 15, 20, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Ajustar altura de filas
        ws.row_dimensions[1].height = 45
        ws.row_dimensions[6].height = 25
        ws.row_dimensions[7].height = 30
        ws.row_dimensions[10].height = 20
        
        wb.save(file_path)
        
        # Actualizar estadísticas del reporte
        report.total_registros = total_productos
        report.monto_total = total_general
        report.save()
        
        print(f"Excel de stock generado exitosamente. Registros: {total_productos}, Total: ${total_general}")
        
        return True
    except Exception as e:
        print(f"Error generando Excel de stock: {e}")
        import traceback
        print(traceback.format_exc())
        return False

@role_required('admin')
def download_report(request, report_id):
    report = get_object_or_404(Report, id=report_id)
    
    # Construir ruta completa del archivo
    if report.ruta_archivo.startswith('reports/'):
        file_path = os.path.join(settings.MEDIA_ROOT, report.ruta_archivo)
    else:
        file_path = report.ruta_archivo
    
    if not os.path.exists(file_path):
        messages.error(request, 'El archivo de reporte no existe')
        return redirect('reportsList', tipo=report.tipo)
    
    try:
        with open(file_path, 'rb') as f:
            content_type = 'application/pdf' if report.formato == 'pdf' else 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response = HttpResponse(f.read(), content_type=content_type)
            filename = os.path.basename(file_path)
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    except Exception as e:
        messages.error(request, f'Error al descargar reporte: {str(e)}')
        return redirect('reportsList', tipo=report.tipo)

@role_required('admin')
def preview_report(request, report_id):
    report = get_object_or_404(Report, id=report_id)

    if report.formato != 'pdf':
        messages.warning(request, 'La vista previa solo está disponible para reportes en PDF')
        return redirect('reportsList', tipo=report.tipo)

    # Construir ruta completa del archivo
    if report.ruta_archivo.startswith('reports/'):
        file_path = os.path.join(settings.MEDIA_ROOT, report.ruta_archivo)
    else:
        file_path = report.ruta_archivo

    if not os.path.exists(file_path):
        messages.error(request, 'El archivo de reporte no existe')
        return redirect('reportsList', tipo=report.tipo)

    try:
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
            return response
    except Exception as e:
        messages.error(request, f'Error al mostrar vista previa: {str(e)}')
        return redirect('reportsList', tipo=report.tipo)

@role_required('admin')
def delete_report(request):
    if request.method == 'POST':
        report_id = request.POST.get('report_id')
        report = get_object_or_404(Report, id=report_id)
        
        try:
            tipo = report.tipo  # Guardamos antes de borrar

            # Construir ruta completa del archivo
            if report.ruta_archivo.startswith('reports/'):
                file_path = os.path.join(settings.MEDIA_ROOT, report.ruta_archivo)
            else:
                file_path = report.ruta_archivo

            if os.path.exists(file_path):
                os.remove(file_path)
            report.delete()
            messages.success(request, 'Reporte eliminado exitosamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar reporte: {str(e)}')

        return redirect('reportsList', tipo=tipo)
    
    return redirect('reports:list')
